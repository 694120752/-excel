
from openpyxl import Workbook
from openpyxl import load_workbook
name = input("请朱瑷悦输入想要处理的文件名：\n")

if( ".xlsx" not in name) :
    name = name + '.xlsx'

wb = Workbook()
dest_filename = name

wb=load_workbook(name)#这是一个已存在的文件
ws=wb['Sheet1'] #获取名为range names的sheet页
row_max = ws.max_row

# ws.cell(row = 2,column=6,value=11)

for i in range(2, row_max+1):

    ws.cell(row=i, column=6).value=ws.cell(row=i,column=4).value * ws.cell(row=i,column=5).value

wb.save(filename=dest_filename)